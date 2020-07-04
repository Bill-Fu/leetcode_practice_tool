# coding=utf-8
import openpyxl

from config import *
from enum import Enum

"""
Format of Record:
Day | Question_Id | Question_Title | Need_Review | Comments | Date
"""


class RecordField(Enum):
    DAY = 1
    QUESTION_ID = 2
    QUESTION_TITLE = 3
    NEED_REVIEW = 4
    COMMENTS = 5
    DATE = 6


class LeetcodeRecord:
    def __init__(self, records_file):
        self.records_file = records_file

    # 增
    def add_record(self, record):
        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        ws.append(record)

        wb.save(self.records_file)
        wb.close()

    # 删
    def delete_record(self, field, value):
        # TODO: 目前还不支持删除数据
        pass

    # 改
    def update_record(self, field, value, record):
        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        for row in range(ws.min_row + 1, ws.max_row + 1):
            if ws.cell(row=row, column=field).value == value:
                for col, content in enumerate(record, 1):
                    ws.cell(row=row, column=col, value=content)

        wb.save(self.records_file)
        wb.close()

    # 查
    def read_record(self, field=None, value=None):
        matched = []

        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        for row in range(ws.min_row + 1, ws.max_row + 1):
            if field is None or ws.cell(row=row, column=field).value == value:
                matched.append([ws.cell(row=row, column=column).value for column in range(ws.min_column, ws.max_column + 1)])

        wb.close()

        return matched


leetcodeRecord = LeetcodeRecord(RECORDS_FILE)