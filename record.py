import openpyxl
import datetime

from config import *

"""
Format of Record:
Day | Question_Id | Question_Title | Need_Review | Comments | Date
"""
class LeetcodeRecord:
    def __init__(self, records_file):
        self.records_file = records_file

    def read_records_file(self):
        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        return [row for row in ws.values][1:]

    def write_record_to_file(self, record):
        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        record.append(datetime.date.today())
        ws.append(record)

        wb.save(self.records_file)

    def update_review_by_id(self, record):
        wb = openpyxl.load_workbook(self.records_file)
        ws = wb.active

        for row in ws.iter_rows():
            if row[1].value == record[0]:
                ws.cell(row=row[1].row, column=4, value=row[3].value + record[1])
                ws.cell(row=row[1].row, column=5, value=row[4].value + "\n" + record[2].decode('utf-8'))
                break

        wb.save(self.records_file)

leetcodeRecord = LeetcodeRecord(RECORDS_FILE)